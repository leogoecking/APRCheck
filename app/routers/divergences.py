from __future__ import annotations

import csv
import io
from io import BytesIO

from fastapi import APIRouter, Depends, Request
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

from app.db import get_db
from app.schemas.forms import DivergenceFilters
from app.services.comparison_service import ensure_latest_competencia_runs, list_divergence_items
from app.utils.web import paginate, pop_flash


router = APIRouter(prefix="/divergences", tags=["divergences"])


def _divergence_sort_state(sort: str | None, direction: str | None) -> tuple[str, str]:
    allowed_sorts = {"competencia", "batch_id", "comparison_id", "apr_id", "categoria", "created_at"}
    safe_sort = sort if sort in allowed_sorts else "created_at"
    safe_direction = "asc" if direction == "asc" else "desc"
    return safe_sort, safe_direction


def build_divergence_export_rows(
    items: list[tuple[object, object, object, dict[str, str]]],
) -> list[list[str]]:
    rows = [
        [
            "apr_id",
            "competencia",
            "assunto",
            "data_abertura",
            "categoria",
            "detalhe",
        ]
    ]
    for item, comparison_run, batch, preview in items:
        rows.append(
            [
                item.apr_id or "",
                comparison_run.competencia,
                preview.get("assunto", ""),
                preview.get("data_abertura", ""),
                item.status_comparacao,
                item.detalhe or "",
            ]
        )
    return rows


def build_divergence_xlsx_rows(
    items: list[tuple[object, object, object, dict[str, str]]],
) -> list[list[str]]:
    rows = [
        [
            "APR ID",
            "Mês",
            "Assunto",
            "Data de abertura",
            "Categoria",
            "Detalhe",
        ]
    ]
    for item, comparison_run, batch, preview in items:
        rows.append(
            [
                item.apr_id or "",
                comparison_run.competencia,
                preview.get("assunto", ""),
                preview.get("data_abertura", ""),
                item.status_comparacao,
                item.detalhe or "",
            ]
        )
    return rows


def build_divergence_xlsx_bytes(
    items: list[tuple[object, object, object, dict[str, str]]],
) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Divergencias"

    rows = build_divergence_xlsx_rows(items)
    for row in rows:
        worksheet.append(row)

    header_fill = PatternFill(fill_type="solid", fgColor="1F6F62")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    widths = {
        "A": 18,
        "B": 14,
        "C": 40,
        "D": 18,
        "E": 24,
        "F": 60,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


@router.get("")
def divergences_page(
    request: Request,
    competencia: str | None = None,
    categoria: str | None = None,
    apr_id: str | None = None,
    sort: str | None = None,
    direction: str | None = None,
    page: int = 1,
    db: Session = Depends(get_db),
) -> object:
    filters = DivergenceFilters(
        competencia=competencia,
        categoria=categoria,
        apr_id=apr_id,
    )
    ensure_latest_competencia_runs(db, competencia=filters.competencia)
    safe_sort, safe_direction = _divergence_sort_state(sort, direction)
    all_items = list_divergence_items(
        db,
        competencia=filters.competencia,
        categoria=filters.categoria,
        apr_id=filters.apr_id,
        sort_by=safe_sort,
        direction=safe_direction,
    )
    pagination = paginate(all_items, page, 20)
    context = {
        "request": request,
        "items": pagination["items"],
        "filters": filters,
        "sort": safe_sort,
        "direction": safe_direction,
        "pagination": pagination,
        "flash": pop_flash(request),
    }
    return request.app.state.templates.TemplateResponse(request, "divergences/index.html", context)


@router.get("/export")
def export_divergences(
    competencia: str | None = None,
    categoria: str | None = None,
    apr_id: str | None = None,
    db: Session = Depends(get_db),
) -> StreamingResponse:
    filters = DivergenceFilters(
        competencia=competencia,
        categoria=categoria,
        apr_id=apr_id,
    )
    ensure_latest_competencia_runs(db, competencia=filters.competencia)
    items = list_divergence_items(
        db,
        competencia=filters.competencia,
        categoria=filters.categoria,
        apr_id=filters.apr_id,
    )
    stream = io.StringIO()
    writer = csv.writer(stream)
    for row in build_divergence_export_rows(items):
        writer.writerow(row)
    response = StreamingResponse(iter([stream.getvalue()]), media_type="text/csv; charset=utf-8")
    response.headers["Content-Disposition"] = 'attachment; filename="divergencias.csv"'
    return response


@router.get("/export.xlsx")
def export_divergences_xlsx(
    competencia: str | None = None,
    categoria: str | None = None,
    apr_id: str | None = None,
    db: Session = Depends(get_db),
) -> StreamingResponse:
    filters = DivergenceFilters(
        competencia=competencia,
        categoria=categoria,
        apr_id=apr_id,
    )
    ensure_latest_competencia_runs(db, competencia=filters.competencia)
    items = list_divergence_items(
        db,
        competencia=filters.competencia,
        categoria=filters.categoria,
        apr_id=filters.apr_id,
    )
    payload = build_divergence_xlsx_bytes(items)
    response = StreamingResponse(
        iter([payload]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response.headers["Content-Disposition"] = 'attachment; filename="divergencias.xlsx"'
    return response
