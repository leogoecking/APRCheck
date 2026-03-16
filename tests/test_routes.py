from __future__ import annotations

from datetime import date, timedelta
from io import BytesIO

from openpyxl import load_workbook
from starlette.datastructures import UploadFile
from starlette.middleware.sessions import SessionMiddleware
from starlette.requests import Request

from app.models.entities import ComparisonRun, ManualAPR, ManualAPRAuditLog
from app.services.comparison_service import list_divergence_items
from app.routers.comparisons import (
    comparison_detail,
    execute_comparison,
    execute_comparison_by_competencia,
)
from app.routers.divergences import (
    build_divergence_export_rows,
    build_divergence_xlsx_bytes,
    divergences_page,
    export_divergences,
    export_divergences_xlsx,
)
from app.routers.history import history_page
from app.routers.imports import import_batch_delete, import_batch_delete_confirm, import_file, imports_page
from app.routers.manual_aprs import (
    manual_apr_create,
    manual_apr_delete,
    manual_apr_delete_confirm,
    manual_apr_edit,
    manual_apr_export_csv,
    manual_apr_import_csv,
    manual_apr_list,
)
from app.services.manual_apr_service import export_manual_aprs_csv_rows


def make_request(app, method: str = "GET", path: str = "/") -> Request:
    return Request(
        {
            "type": "http",
            "http_version": "1.1",
            "method": method,
            "scheme": "http",
            "path": path,
            "raw_path": path.encode(),
            "query_string": b"",
            "headers": [],
            "client": ("testclient", 50000),
            "server": ("testserver", 80),
            "app": app,
        }
    )



def test_manual_apr_create_and_duplicate(app_module):
    db = app_module.db_module.SessionLocal()
    try:
        response = manual_apr_create(
            make_request(app_module.app, method="POST", path="/manual-aprs"),
            apr_id="APR-10",
            data_referencia="2026-03-01",
            responsavel="Maria",
            descricao=None,
            observacao=None,
            status_apr="ativo",
            db=db,
        )
        assert response.status_code == 303

        duplicate = manual_apr_create(
            make_request(app_module.app, method="POST", path="/manual-aprs"),
            apr_id="APR-10",
            data_referencia="2026-03-01",
            responsavel="Joao",
            descricao=None,
            observacao=None,
            status_apr="ativo",
            db=db,
        )
        assert duplicate.status_code == 409
        assert "já existe" in duplicate.body.decode()
    finally:
        db.close()


def test_manual_apr_list_shows_month_recognition(app_module):
    db = app_module.db_module.SessionLocal()
    try:
        today = date.today()
        current_month_date = today.replace(day=min(today.day, 15))
        previous_month_date = (today.replace(day=1) - timedelta(days=1)).replace(day=10)

        first_response = manual_apr_create(
            make_request(app_module.app, method="POST", path="/manual-aprs"),
            apr_id="APR-30",
            data_referencia=current_month_date.isoformat(),
            responsavel="Maria",
            descricao="MANUTENCAO A",
            observacao=None,
            status_apr="ativo",
            db=db,
        )
        second_response = manual_apr_create(
            make_request(app_module.app, method="POST", path="/manual-aprs"),
            apr_id="APR-20",
            data_referencia=previous_month_date.isoformat(),
            responsavel="Joao",
            descricao="MANUTENCAO B",
            observacao=None,
            status_apr="ativo",
            db=db,
        )
        assert first_response.status_code == 303
        assert second_response.status_code == 303

        page = manual_apr_list(
            make_request(app_module.app, path="/manual-aprs"),
            q=None,
            sort="apr_id",
            direction="asc",
            page=1,
            db=db,
        )
        body = page.body.decode()
        assert page.status_code == 200
        assert "Mês atual" in body
        assert "Mês anterior" in body
        assert "Colaborador" in body
        assert "MANUTENCAO A" in body
        assert "MANUTENCAO B" in body
        assert body.index("APR-20") < body.index("APR-30")
    finally:
        db.close()


def test_manual_apr_csv_import_and_export(app_module):
    db = app_module.db_module.SessionLocal()
    try:
        import_response = manual_apr_import_csv(
            make_request(app_module.app, method="POST", path="/manual-aprs/import-csv"),
            arquivo=UploadFile(
                filename="base-manual.csv",
                file=BytesIO(
                    b"apr_id,data_abertura,assunto,colaborador\nAPR-40,2026-03-11,ASSUNTO CSV,Ana\n"
                ),
            ),
            db=db,
        )
        assert import_response.status_code == 303

        export_response = manual_apr_export_csv(q=None, sort="apr_id", direction="asc", db=db)
        assert export_response.status_code == 200
        assert "text/csv" in export_response.headers["content-type"]
        assert "base_manual_aprs.csv" in export_response.headers["content-disposition"]
        exported_rows = export_manual_aprs_csv_rows(db.query(ManualAPR).order_by(ManualAPR.apr_id.asc()).all())
        assert exported_rows[0] == [
            "apr_id",
            "data_abertura",
            "assunto",
            "colaborador",
            "created_at",
            "updated_at",
        ]
    finally:
        db.close()


def test_manual_apr_delete_requires_confirmation_and_reruns_comparison(app_module):
    db = app_module.db_module.SessionLocal()
    try:
        create_response = manual_apr_create(
            make_request(app_module.app, method="POST", path="/manual-aprs"),
            apr_id="APR-55",
            data_referencia="2026-03-11",
            responsavel="Maria",
            descricao="ASSUNTO",
            observacao=None,
            status_apr="ativo",
            db=db,
        )
        assert create_response.status_code == 303

        import_response = import_file(
            make_request(app_module.app, method="POST", path="/imports"),
            competencia="2026-03",
            arquivo=UploadFile(
                filename="lote.csv",
                file=BytesIO(b"apr_id,descricao\nAPR-55,Conciliado\n"),
            ),
            db=db,
        )
        assert import_response.status_code == 303
        execute_comparison_by_competencia(
            make_request(app_module.app, method="POST", path="/comparisons/run-by-competencia"),
            competencia="2026-03",
            db=db,
        )

        confirm_page = manual_apr_delete_confirm(
            make_request(app_module.app, path="/manual-aprs/1/delete"),
            manual_apr_id=1,
            db=db,
        )
        assert confirm_page.status_code == 200
        assert "Digite o ID para confirmar" in confirm_page.body.decode()

        invalid_delete = manual_apr_delete(
            make_request(app_module.app, method="POST", path="/manual-aprs/1/delete"),
            manual_apr_id=1,
            confirm_apr_id="ERRADO",
            db=db,
        )
        assert invalid_delete.status_code == 400

        valid_delete = manual_apr_delete(
            make_request(app_module.app, method="POST", path="/manual-aprs/1/delete"),
            manual_apr_id=1,
            confirm_apr_id="APR-55",
            db=db,
        )
        assert valid_delete.status_code == 303

        runs = list(
            db.query(ComparisonRun)
            .filter(ComparisonRun.scope_type == "competencia", ComparisonRun.competencia == "2026-03")
            .order_by(ComparisonRun.id.asc())
        )
        delete_audit = db.query(ManualAPRAuditLog).filter_by(action="delete", apr_id="APR-55").one()
        assert len(runs) == 2
        assert runs[-1].total_manual == 0
        assert runs[-1].total_conciliado == 0
        assert runs[-1].total_faltando_manual == 1
        assert "apr_id=APR-55" in (delete_audit.detalhe or "")
    finally:
        db.close()


def test_manual_apr_edit_keeps_hidden_legacy_fields(app_module):
    db = app_module.db_module.SessionLocal()
    try:
        create_response = manual_apr_create(
            make_request(app_module.app, method="POST", path="/manual-aprs"),
            apr_id="APR-EDIT-1",
            data_referencia="2026-03-11",
            responsavel="Maria",
            descricao="ASSUNTO ANTIGO",
            observacao="OBS LEGADA",
            status_apr="ativo",
            db=db,
        )
        assert create_response.status_code == 303

        edit_response = manual_apr_edit(
            make_request(app_module.app, method="POST", path="/manual-aprs/1/edit"),
            manual_apr_id=1,
            apr_id="APR-EDIT-1",
            data_referencia="2026-03-12",
            responsavel="Carlos",
            descricao="ASSUNTO NOVO",
            observacao=None,
            status_apr=None,
            db=db,
        )
        assert edit_response.status_code == 303

        manual = db.query(ManualAPRAuditLog).filter_by(action="update", apr_id="APR-EDIT-1").one()
        assert "colaborador=Carlos" in (manual.detalhe or "")
        assert "assunto=ASSUNTO NOVO" in (manual.detalhe or "")

        saved = db.query(ManualAPR).filter_by(apr_id="APR-EDIT-1").one()
        assert saved.status == "ativo"
        assert saved.observacao == "OBS LEGADA"
    finally:
        db.close()


def test_import_run_comparison_and_export(app_module):
    db = app_module.db_module.SessionLocal()
    try:
        manual_apr_create(
            make_request(app_module.app, method="POST", path="/manual-aprs"),
            apr_id="APR-1",
            data_referencia="2026-03-01",
            responsavel="Equipe",
            descricao=None,
            observacao=None,
            status_apr="ativo",
            db=db,
        )
        manual_apr_create(
            make_request(app_module.app, method="POST", path="/manual-aprs"),
            apr_id="APR-9",
            data_referencia="2026-03-02",
            responsavel="Equipe",
            descricao=None,
            observacao=None,
            status_apr="ativo",
            db=db,
        )

        upload = UploadFile(
            filename="lote.csv",
            file=BytesIO(
                b"apr_id,descricao\nAPR-1,Conciliado\nAPR-2,Faltando manual\nAPR-2,Duplicado\nAPR-3,Faltando manual valido\n,Invalido\n"
            ),
        )
        response = import_file(
            make_request(app_module.app, method="POST", path="/imports"),
            competencia="2026-03",
            arquivo=upload,
            db=db,
        )
        assert response.status_code == 303
        assert "batch_id=1" in response.headers["location"]

        comparison = execute_comparison(
            make_request(app_module.app, method="POST", path="/comparisons/run/1"),
            batch_id=1,
            db=db,
        )
        assert comparison.status_code == 303
        assert comparison.headers["location"] == "/comparisons/1"
        execute_comparison_by_competencia(
            make_request(app_module.app, method="POST", path="/comparisons/run-by-competencia"),
            competencia="2026-03",
            db=db,
        )

        detail = comparison_detail(
            make_request(app_module.app, path="/comparisons/1"),
            run_id=1,
            db=db,
        )
        assert detail.status_code == 200
        detail_body = detail.body.decode()
        assert "faltando_no_manual" in detail_body
        assert "faltando_no_importado" in detail_body
        assert "duplicado" in detail_body
        assert "invalido" in detail_body

        divergences = divergences_page(
            make_request(app_module.app, path="/divergences"),
            competencia="2026-03",
            db=db,
        )
        assert divergences.status_code == 200
        divergences_body = divergences.body.decode()
        assert "APR-1" not in divergences_body
        assert "Buscar manual" in divergences_body
        assert "Abrir comparação" in divergences_body
        assert "Mês:" not in divergences_body
        assert "Exportar XLSX" in divergences_body

        export = export_divergences(competencia="2026-03", db=db)
        assert export.status_code == 200
        assert "text/csv" in export.headers["content-type"]
        assert "divergencias.csv" in export.headers["content-disposition"]
        export_rows = build_divergence_export_rows(
            list_divergence_items(db, competencia="2026-03")
        )
        assert export_rows[0] == [
            "apr_id",
            "competencia",
            "assunto",
            "data_abertura",
            "categoria",
            "detalhe",
        ]

        xlsx_export = export_divergences_xlsx(competencia="2026-03", db=db)
        assert xlsx_export.status_code == 200
        assert (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            in xlsx_export.headers["content-type"]
        )
        assert "divergencias.xlsx" in xlsx_export.headers["content-disposition"]

        workbook = load_workbook(
            BytesIO(build_divergence_xlsx_bytes(list_divergence_items(db, competencia="2026-03")))
        )
        worksheet = workbook.active
        assert [cell.value for cell in worksheet[1]] == [
            "APR ID",
            "Mês",
            "Assunto",
            "Data de abertura",
            "Categoria",
            "Detalhe",
        ]
        exported_rows = list(worksheet.iter_rows(min_row=2, values_only=True))
        apr_3_row = next(
            row for row in exported_rows if row[0] == "APR-3" and row[4] == "faltando_no_manual"
        )
        assert apr_3_row[1] == "2026-03"
        assert apr_3_row[2] == "Faltando manual valido"
        assert apr_3_row[4] == "faltando_no_manual"
        assert apr_3_row[5] == "ID encontrado no escopo importado e ausente no cadastro manual da competência."
    finally:
        db.close()


def test_divergences_prioritize_apr_id_and_month_view(app_module):
    db = app_module.db_module.SessionLocal()
    try:
        manual_apr_create(
            make_request(app_module.app, method="POST", path="/manual-aprs"),
            apr_id="APR-1",
            data_referencia="2026-03-01",
            responsavel="Equipe",
            descricao=None,
            observacao=None,
            status_apr="ativo",
            db=db,
        )

        for filename, content in (
            ("lote-a.csv", b"apr_id,descricao\nAPR-1,Conciliado\nAPR-2,Faltando manual\n"),
            ("lote-b.csv", b"apr_id,descricao\nAPR-3,Faltando manual\n"),
        ):
            response = import_file(
                make_request(app_module.app, method="POST", path="/imports"),
                competencia="2026-03",
                arquivo=UploadFile(filename=filename, file=BytesIO(content)),
                db=db,
            )
            assert response.status_code == 303

        execute_comparison(
            make_request(app_module.app, method="POST", path="/comparisons/run/1"),
            batch_id=1,
            db=db,
        )
        execute_comparison(
            make_request(app_module.app, method="POST", path="/comparisons/run/2"),
            batch_id=2,
            db=db,
        )
        execute_comparison_by_competencia(
            make_request(app_module.app, method="POST", path="/comparisons/run-by-competencia"),
            competencia="2026-03",
            db=db,
        )

        apr_filtered = divergences_page(
            make_request(app_module.app, path="/divergences"),
            competencia="2026-03",
            apr_id="APR-2",
            db=db,
        )
        apr_body = apr_filtered.body.decode()
        assert apr_filtered.status_code == 200
        assert "APR-2" in apr_body
        assert "APR-3" not in apr_body
        assert "Mês" in apr_body
        assert "Lote" not in apr_body

        month_filtered = divergences_page(
            make_request(app_module.app, path="/divergences"),
            competencia="2026-03",
            db=db,
        )
        month_body = month_filtered.body.decode()
        assert month_filtered.status_code == 200
        assert "2026-03" in month_body
        assert "Escopo" not in month_body

        export = export_divergences(
            competencia="2026-03",
            apr_id="APR-2",
            db=db,
        )
        assert export.status_code == 200
        assert "text/csv" in export.headers["content-type"]
        assert "divergencias.csv" in export.headers["content-disposition"]
    finally:
        db.close()


def test_import_batch_delete_requires_confirmation_and_updates_remaining_competencia_runs(app_module):
    db = app_module.db_module.SessionLocal()
    try:
        manual_apr_create(
            make_request(app_module.app, method="POST", path="/manual-aprs"),
            apr_id="APR-1",
            data_referencia="2026-03-01",
            responsavel="Equipe",
            descricao=None,
            observacao=None,
            status_apr="ativo",
            db=db,
        )

        for filename, content in (
            ("lote-a.csv", b"apr_id,descricao\nAPR-1,Conciliado\nAPR-2,Faltando manual\n"),
            ("lote-b.csv", b"apr_id,descricao\nAPR-3,Novo\n"),
        ):
            response = import_file(
                make_request(app_module.app, method="POST", path="/imports"),
                competencia="2026-03",
                arquivo=UploadFile(filename=filename, file=BytesIO(content)),
                db=db,
            )
            assert response.status_code == 303

        execute_comparison(
            make_request(app_module.app, method="POST", path="/comparisons/run/1"),
            batch_id=1,
            db=db,
        )
        execute_comparison(
            make_request(app_module.app, method="POST", path="/comparisons/run/2"),
            batch_id=2,
            db=db,
        )
        execute_comparison_by_competencia(
            make_request(app_module.app, method="POST", path="/comparisons/run-by-competencia"),
            competencia="2026-03",
            db=db,
        )

        confirm_page = import_batch_delete_confirm(
            make_request(app_module.app, path="/imports/1/delete"),
            batch_id=1,
            db=db,
        )
        assert confirm_page.status_code == 200
        assert "Digite o ID do lote para confirmar" in confirm_page.body.decode()

        invalid_delete = import_batch_delete(
            make_request(app_module.app, method="POST", path="/imports/1/delete"),
            batch_id=1,
            confirm_batch_id="999",
            db=db,
        )
        assert invalid_delete.status_code == 400

        valid_delete = import_batch_delete(
            make_request(app_module.app, method="POST", path="/imports/1/delete"),
            batch_id=1,
            confirm_batch_id="1",
            db=db,
        )
        assert valid_delete.status_code == 303

        remaining_batches = list(
            db.query(ComparisonRun).filter(ComparisonRun.batch_id == 2).order_by(ComparisonRun.id.asc())
        )
        deleted_batch_runs = list(
            db.query(ComparisonRun).filter(ComparisonRun.batch_id == 1).order_by(ComparisonRun.id.asc())
        )
        assert len(deleted_batch_runs) == 1
        assert deleted_batch_runs[0].scope_type == "batch"
        assert remaining_batches[-1].total_manual == 1
        assert remaining_batches[-1].total_faltando_importado == 1
        imports_page_response = imports_page(
            make_request(app_module.app, path="/imports"),
            db=db,
        )
        assert imports_page_response.status_code == 200
        assert "lote-a.csv" not in imports_page_response.body.decode()
    finally:
        db.close()


def test_divergences_and_history_support_pagination(app_module):
    db = app_module.db_module.SessionLocal()
    try:
        manual_apr_create(
            make_request(app_module.app, method="POST", path="/manual-aprs"),
            apr_id="APR-1",
            data_referencia="2026-03-01",
            responsavel="Equipe",
            descricao=None,
            observacao=None,
            status_apr="ativo",
            db=db,
        )

        for index in range(25):
            response = import_file(
                make_request(app_module.app, method="POST", path="/imports"),
                competencia="2026-03",
                arquivo=UploadFile(
                    filename=f"lote-{index}.csv",
                    file=BytesIO(
                        f"ID,Assunto,Abertura\nAPR-X-{index},Assunto {index},11/03/2026 13:10\n".encode("utf-8")
                    ),
                ),
                db=db,
            )
            assert response.status_code == 303
            execute_comparison(
                make_request(app_module.app, method="POST", path=f"/comparisons/run/{index + 1}"),
                batch_id=index + 1,
                db=db,
            )
        execute_comparison_by_competencia(
            make_request(app_module.app, method="POST", path="/comparisons/run-by-competencia"),
            competencia="2026-03",
            db=db,
        )

        divergences = divergences_page(
            make_request(app_module.app, path="/divergences"),
            competencia="2026-03",
            page=2,
            db=db,
        )
        assert divergences.status_code == 200
        divergences_body = divergences.body.decode()
        assert "Página 2 de" in divergences_body
        assert "Assunto" in divergences_body
        assert "Data de abertura" in divergences_body
        assert "11/03/2026" in divergences_body

        history = history_page(
            make_request(app_module.app, path="/history"),
            import_page=2,
            comparison_page=2,
            audit_page=1,
            db=db,
        )
        assert history.status_code == 200
        history_body = history.body.decode()
        assert "Página 2 de" in history_body
        assert "Auditoria Manual" in history_body
    finally:
        db.close()


def test_manual_apr_create_reruns_existing_comparisons(app_module):
    db = app_module.db_module.SessionLocal()
    try:
        upload = UploadFile(
            filename="lote.csv",
            file=BytesIO(b"apr_id,descricao\nAPR-1,Conciliado\nAPR-2,Faltando manual\n"),
        )
        import_response = import_file(
            make_request(app_module.app, method="POST", path="/imports"),
            competencia="2026-03",
            arquivo=upload,
            db=db,
        )
        assert import_response.status_code == 303

        comparison_response = execute_comparison(
            make_request(app_module.app, method="POST", path="/comparisons/run/1"),
            batch_id=1,
            db=db,
        )
        assert comparison_response.status_code == 303

        monthly_response = execute_comparison_by_competencia(
            make_request(app_module.app, method="POST", path="/comparisons/run-by-competencia"),
            competencia="2026-03",
            db=db,
        )
        assert monthly_response.status_code == 303

        run_before = (
            db.query(ComparisonRun)
            .filter(ComparisonRun.scope_type == "competencia", ComparisonRun.competencia == "2026-03")
            .order_by(ComparisonRun.id.asc())
            .first()
        )
        assert run_before.total_manual == 0
        assert run_before.total_conciliado == 0
        assert run_before.total_faltando_manual == 2
        assert run_before.scope_type == "competencia"

        create_response = manual_apr_create(
            make_request(app_module.app, method="POST", path="/manual-aprs"),
            apr_id="APR-1",
            data_referencia="2026-03-10",
            responsavel="Maria",
            descricao=None,
            observacao=None,
            status_apr="ativo",
            db=db,
        )
        assert create_response.status_code == 303

        runs = list(
            db.query(ComparisonRun)
            .filter(ComparisonRun.scope_type == "competencia", ComparisonRun.competencia == "2026-03")
            .order_by(ComparisonRun.id.asc())
        )
        assert len(runs) == 2
        run_after = runs[-1]
        assert run_after.total_manual == 1
        assert run_after.total_conciliado == 1
        assert run_after.total_faltando_manual == 1
    finally:
        db.close()


def test_flash_message_is_rendered_after_redirect(app_module):
    db = app_module.db_module.SessionLocal()
    try:
        assert any(middleware.cls is SessionMiddleware for middleware in app_module.app.user_middleware)

        request = make_request(app_module.app, method="POST", path="/manual-aprs")
        request.scope["session"] = {}
        response = manual_apr_create(
            request,
            apr_id="APR-FLASH-1",
            data_referencia="2026-03-10",
            responsavel="Maria",
            descricao="Teste flash",
            observacao="",
            status_apr="ativo",
            db=db,
        )

        assert response.status_code == 303
        assert request.scope["session"]["_flash"]["message"] == (
            "APR manual cadastrada com sucesso. Referência reconhecida: mês atual."
        )

        listing_request = make_request(app_module.app, path="/manual-aprs")
        listing_request.scope["session"] = request.scope["session"]
        listing = manual_apr_list(listing_request, db=db)

        assert listing.status_code == 200
        assert "APR manual cadastrada com sucesso. Referência reconhecida: mês atual." in listing.body.decode()
    finally:
        db.close()


def test_execute_comparison_by_competencia_creates_consolidated_run(app_module):
    db = app_module.db_module.SessionLocal()
    try:
        manual_apr_create(
            make_request(app_module.app, method="POST", path="/manual-aprs"),
            apr_id="APR-1",
            data_referencia="2026-03-01",
            responsavel="Equipe",
            descricao=None,
            observacao=None,
            status_apr="ativo",
            db=db,
        )

        for filename, content in (
            ("lote-a.csv", b"apr_id,descricao\nAPR-1,Conciliado\nAPR-2,Faltando manual\n"),
            ("lote-b.csv", b"apr_id,descricao\nAPR-2,Duplicado em lote diferente\nAPR-3,Novo\n"),
        ):
            response = import_file(
                make_request(app_module.app, method="POST", path="/imports"),
                competencia="2026-03",
                arquivo=UploadFile(filename=filename, file=BytesIO(content)),
                db=db,
            )
            assert response.status_code == 303

        comparison_response = execute_comparison_by_competencia(
            make_request(app_module.app, method="POST", path="/comparisons/run-by-competencia"),
            competencia="2026-03",
            db=db,
        )
        assert comparison_response.status_code == 303
        assert comparison_response.headers["location"] == "/comparisons/1"

        run = db.query(ComparisonRun).filter(ComparisonRun.id == 1).one()
        assert run.scope_type == "competencia"
        assert run.total_conciliado == 1
        assert run.total_faltando_manual == 1
        assert run.total_duplicados == 1
    finally:
        db.close()
